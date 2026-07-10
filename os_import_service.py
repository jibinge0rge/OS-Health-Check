from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}


def _extension(filename: str) -> str:
    return Path(filename or "").suffix.lower()


def assert_supported_filename(filename: str) -> str:
    extension = _extension(filename)
    if extension not in SUPPORTED_EXTENSIONS:
        raise ValueError("Only CSV (.csv) and Excel (.xlsx) files are supported.")
    return extension


def _normalize_headers(raw_headers: list[object]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(raw_headers):
        label = str(value or "").strip() or f"Column {index + 1}"
        count = seen.get(label, 0) + 1
        seen[label] = count
        headers.append(label if count == 1 else f"{label} ({count})")
    return headers


def _read_csv_table(content: bytes) -> tuple[list[str], list[list[str]]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Could not decode CSV file.")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = [[str(cell or "").strip() for cell in row] for row in reader]
    rows = [row for row in rows if any(cell for cell in row)]
    if not rows:
        raise ValueError("File has no rows.")

    width = max(len(row) for row in rows)
    normalized_rows = [row + [""] * (width - len(row)) for row in rows]
    headers = _normalize_headers(normalized_rows[0])
    data_rows = normalized_rows[1:]
    return headers, data_rows


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_xlsx_table(content: bytes) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        raw_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = [_cell_to_text(cell) for cell in row]
            if any(values):
                raw_rows.append(values)
    finally:
        workbook.close()

    if not raw_rows:
        raise ValueError("Excel sheet has no rows.")

    width = max(len(row) for row in raw_rows)
    normalized_rows = [row + [""] * (width - len(row)) for row in raw_rows]
    headers = _normalize_headers(normalized_rows[0])
    data_rows = normalized_rows[1:]
    return headers, data_rows


def read_tabular_file(content: bytes, filename: str) -> tuple[list[str], list[list[str]]]:
    extension = assert_supported_filename(filename)
    if not content:
        raise ValueError("Uploaded file is empty.")
    if extension == ".csv":
        return _read_csv_table(content)
    return _read_xlsx_table(content)


def inspect_os_import_file(content: bytes, filename: str) -> dict[str, object]:
    headers, data_rows = read_tabular_file(content, filename)
    return {
        "filename": Path(filename).name,
        "columns": headers,
        "row_count": len(data_rows),
    }


def extract_distinct_os_values(
    content: bytes,
    filename: str,
    columns: list[str],
) -> dict[str, object]:
    selected = [str(column or "").strip() for column in columns if str(column or "").strip()]
    if not selected:
        raise ValueError("Select at least one column.")

    headers, data_rows = read_tabular_file(content, filename)
    header_indexes = {header: index for index, header in enumerate(headers)}
    missing = [column for column in selected if column not in header_indexes]
    if missing:
        raise ValueError(f"Unknown column(s): {', '.join(missing)}")

    indexes = [header_indexes[column] for column in selected]
    seen: set[str] = set()
    values: list[str] = []
    for row in data_rows:
        for index in indexes:
            value = row[index].strip() if index < len(row) else ""
            if not value:
                continue
            key = " ".join(value.lower().split())
            if key in seen:
                continue
            seen.add(key)
            values.append(value)

    return {
        "filename": Path(filename).name,
        "columns": selected,
        "values": values,
        "distinct_count": len(values),
        "row_count": len(data_rows),
    }
